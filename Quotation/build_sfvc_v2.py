import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding='utf-8')


# ─── PATHS ───────────────────────────────────────────────────────────────────
base_dir = r"d:\Project\Sunford Valves\Quotation"
input_xlsx = os.path.join(base_dir, "SFVC_Product_Catalog.xlsx")
output_xlsx = os.path.join(base_dir, "SFVC_Product_Catalog_v2.xlsx")

# ─── LOAD DATA ───────────────────────────────────────────────────────────────
# Load clean data from '📋 Product Manager' sheet (skipping first 6 header rows)
df = pd.read_excel(input_xlsx, sheet_name='📋 Product Manager', skiprows=6)

# Rename columns to match expected script format
df = df.rename(columns={
    'Rate (INR)': 'Rate',
    'Design_Standard': 'Design_Std',
    'Hydro_Test': 'Hydro_Test'
})

# Parse design standard & hydro from description once
def parse_desc_fields(desc_str):
    parts = [p.strip() for p in str(desc_str).split('|') if p.strip()]
    design_std, hydro, ff, pressure = '', '', '', ''
    for p in parts:
        if p.startswith('Design:'): design_std = p.replace('Design:','').strip()
        elif p.startswith('Hydro:'): hydro = p.replace('Hydro:','').strip()
        elif p.startswith('F-F:'): ff = p.replace('F-F:','').strip()
        elif p.startswith('Pressure:'): pressure = p.replace('Pressure:','').strip()
    return design_std, hydro, ff, pressure

# Fill in Design_Std, Hydro_Test, FF, Pressure_Std
parsed_design = df['Description'].apply(lambda x: parse_desc_fields(x)[0])
parsed_hydro = df['Description'].apply(lambda x: parse_desc_fields(x)[1])

df['Design_Std'] = df['Design_Std'].fillna(parsed_design)
df['Hydro_Test'] = df['Hydro_Test'].fillna(parsed_hydro)
df['FF'] = df['Description'].apply(lambda x: parse_desc_fields(x)[2])
df['Pressure_Std'] = df['Description'].apply(lambda x: parse_desc_fields(x)[3])

# ─── COLORS ──────────────────────────────────────────────────────────────────
CLR = {
    'navy':     'FF1A3A5C',
    'steel':    'FF2E6B9E',
    'sky':      'FFD0E4F5',
    'sky2':     'FFE8F4FF',
    'gold':     'FFC9A84C',
    'gold_lt':  'FFFEF3D0',
    'white':    'FFFFFFFF',
    'gray_lt':  'FFF5F5F5',
    'gray_bd':  'FFB0B0B0',
    'green':    'FF1B7A1B',
    'green_lt': 'FFEBF9EB',
    'red':      'FFCC2929',
    'red_lt':   'FFFDE8E8',
    'on_req':   'FF7B4FA6',
    'on_req_lt':'FFF3ECFA',
    'dark_row': 'FF0D2137',
    'black':    'FF000000',
}

TAB_COLORS = {
    'Ball Valves':       'FF1565C0',
    'Gate Valves':       'FF2E7D32',
    'Globe Valves':      'FF6A1B9A',
    'Check Valves':      'FFE65100',
    'Control Valves':    'FF00838F',
    'Butterfly Valves':  'FFC62828',
    'Strainers':         'FF4527A0',
    'Piston Valves':     'FF00695C',
    'Safety Valves':     'FFAD1457',
}

# ─── BORDER HELPERS ──────────────────────────────────────────────────────────
def thin():
    t = Side(style='thin', color='FFB0B0B0')
    return Border(left=t, right=t, top=t, bottom=t)

def set_widths(ws, wd):
    for col, w in wd.items():
        ws.column_dimensions[get_column_letter(col)].width = w

# ─── MASTER SHEET COLUMN MAP ──────────────────────────────────────────────────
PM = '\'📋 Product Manager\''  # escaped sheet name for formulas

# ─── WORKBOOK ─────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: PRODUCT MANAGER (Single Source of Truth)
# ══════════════════════════════════════════════════════════════════════════════
ws_pm = wb.create_sheet('📋 Product Manager', 0)
ws_pm.sheet_properties.tabColor = '1A3A5C'
ws_pm.freeze_panes = 'A8'

# Row 1: Title
ws_pm.merge_cells('A1:N1')
c = ws_pm['A1']
c.value = 'SUNFORD VALVES COMPANY — PRODUCT MANAGER  (Single Source of Truth)'
c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=13)
c.fill = PatternFill('solid', start_color=CLR['navy'])
c.alignment = Alignment(horizontal='center', vertical='center')
ws_pm.row_dimensions[1].height = 28

# Row 2: Subtitle
ws_pm.merge_cells('A2:N2')
c = ws_pm['A2']
c.value = 'Edit this table only. All type sheets and the price summary auto-refresh from here. Do NOT edit type sheets directly.'
c.font = Font(name='Arial', italic=True, color='FFB0C4DE', size=8)
c.fill = PatternFill('solid', start_color='FF0D2137')
c.alignment = Alignment(horizontal='center', vertical='center')
ws_pm.row_dimensions[2].height = 14

# Rows 3–6: Instruction tiles
instructions = [
    ('🟢 ADD NEW PRODUCT',    'Scroll to the bottom of the table → type a new row. Fill Valve_Type, Valve_Name, MOC, Size, Rate, Status. The type sheet will auto-include it IF Valve_Type matches exactly.',  'FFEBF9EB', 'FF1B7A1B'),
    ('✏️ EDIT PRICE / FIELD', 'Click the cell directly. Change the value. Save (Ctrl+S). All linked sheets will recalculate on next open.',  'FFFFF8E8', 'FF8B4000'),
    ('🔴 HIDE A PRODUCT',     'Set Status column = Inactive. The product stays in data but is marked for the website to filter out.',  'FFFDE8E8', 'FFCC2929'),
    ('🗑️ DELETE A PRODUCT',  'Select the entire row → Right-click → Delete Row. All type sheet formulas will auto-adjust (they use row-specific references).',  'FFF3ECFA', 'FF7B4FA6'),
]
for i, (act, desc, bg, fc) in enumerate(instructions):
    r = 3 + i
    ws_pm.merge_cells(f'A{r}:B{r}')
    c1 = ws_pm.cell(row=r, column=1, value=act)
    c1.font = Font(name='Arial', bold=True, color=fc, size=8)
    c1.fill = PatternFill('solid', start_color=bg)
    c1.alignment = Alignment(horizontal='center', vertical='center')
    c1.border = thin()
    ws_pm.merge_cells(f'C{r}:N{r}')
    c2 = ws_pm.cell(row=r, column=3, value=desc)
    c2.font = Font(name='Arial', size=8, color='FF333333')
    c2.fill = PatternFill('solid', start_color=bg)
    c2.alignment = Alignment(horizontal='left', vertical='center')
    c2.border = thin()
    ws_pm.row_dimensions[r].height = 17

# Row 7: Column headers
PM_COLS = ['Valve_Type','Valve_Name','Description','MOC','Size',
           'Pressure_Class','End_Connection','Design_Standard',
           'Hydro_Test','Unit','Rate_INR','Status','Last_Updated','Remarks']
PM_WIDTHS = {1:16,2:52,3:50,4:34,5:9,6:14,7:18,8:32,9:22,10:7,11:13,12:10,13:14,14:22}
set_widths(ws_pm, PM_WIDTHS)

for ci, h in enumerate(PM_COLS, 1):
    c = ws_pm.cell(row=7, column=ci, value=h)
    c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=9)
    c.fill = PatternFill('solid', start_color=CLR['navy'])
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin()
ws_pm.row_dimensions[7].height = 30

# Data rows (row 8 onward)
PM_DATA_START = 8

def rate_numeric(r):
    try: return float(str(r).strip())
    except: return None

for i, (_, row_data) in enumerate(df.iterrows()):
    r = PM_DATA_START + i
    bg = CLR['white'] if i % 2 == 0 else CLR['gray_lt']
    tab_c = TAB_COLORS.get(str(row_data['Valve_Type']), CLR['steel'])

    vals = [
        row_data['Valve_Type'],
        row_data['Valve_Name'],
        row_data['Description'],
        row_data['MOC'],
        str(row_data['Size']),
        str(row_data.get('Pressure_Class','') or '').replace('nan',''),
        str(row_data.get('End_Connection','') or '').replace('nan',''),
        row_data['Design_Std'],
        row_data['Hydro_Test'],
        str(row_data.get('Unit','Nos') or 'Nos').replace('nan','Nos'),
        row_data['Rate'],
        str(row_data.get('Status','Active') or 'Active').replace('nan','Active'),
        '',
        str(row_data.get('Remarks','') or '').replace('nan',''),
    ]

    for ci, val in enumerate(vals, 1):
        c = ws_pm.cell(row=r, column=ci)
        v = val
        if str(v) == 'nan': v = ''

        if ci == 11:  # Rate
            rn = rate_numeric(v)
            if rn is not None:
                c.value = rn
                c.number_format = '₹\\ #,##0'
                c.font = Font(name='Arial', bold=True, size=9, color='FF000000')
                c.fill = PatternFill('solid', start_color=CLR['gold_lt'])
            else:
                c.value = 'On Request'
                c.font = Font(name='Arial', bold=True, size=9, color=CLR['on_req'][2:])
                c.fill = PatternFill('solid', start_color=CLR['on_req_lt'])
            c.alignment = Alignment(horizontal='right', vertical='center')
        elif ci == 12:  # Status
            c.value = v
            s_bg = CLR['green_lt'] if v == 'Active' else CLR['red_lt']
            s_fc = CLR['green'] if v == 'Active' else CLR['red']
            c.font = Font(name='Arial', bold=True, size=8, color=s_fc)
            c.fill = PatternFill('solid', start_color=s_bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
        elif ci == 1:  # Valve_Type
            c.value = v
            c.font = Font(name='Arial', bold=True, size=8, color=tab_c[2:])
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='left', vertical='center')
        elif ci in (2,3,4):  # long text
            c.value = v
            c.font = Font(name='Arial', size=8)
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        elif ci == 5:  # Size
            c.value = v
            c.font = Font(name='Arial', bold=True, size=9, color='FF1A3A5C')
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
        else:
            c.value = v
            c.font = Font(name='Arial', size=8)
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        c.border = thin()
        ws_pm.row_dimensions[r].height = 15

TOTAL_DATA_ROWS = len(df)
LAST_DATA_ROW = PM_DATA_START + TOTAL_DATA_ROWS - 1

# Dropdown validations
dv_type = DataValidation(
    type='list',
    formula1='"Ball Valves,Gate Valves,Globe Valves,Check Valves,Control Valves,Butterfly Valves,Strainers,Piston Valves,Safety Valves"',
    allow_blank=True, showErrorMessage=True,
    errorTitle='Invalid Type', error='Choose from the list.'
)
ws_pm.add_data_validation(dv_type)
dv_type.sqref = f'A{PM_DATA_START}:A{LAST_DATA_ROW+200}'

dv_status = DataValidation(
    type='list', formula1='"Active,Inactive"',
    allow_blank=True, showErrorMessage=True,
    errorTitle='Invalid Status', error='Choose Active or Inactive.'
)
ws_pm.add_data_validation(dv_status)
dv_status.sqref = f'L{PM_DATA_START}:L{LAST_DATA_ROW+200}'

ws_pm.print_title_rows = '1:7'
ws_pm.page_setup.orientation = 'landscape'
ws_pm.page_setup.fitToPage = True
ws_pm.page_setup.fitToWidth = 1

print("✓ Product Manager sheet built")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: PRICE SUMMARY (formula-driven from Product Manager)
# ══════════════════════════════════════════════════════════════════════════════
ws_ps = wb.create_sheet('📊 Price Summary', 1)
ws_ps.sheet_properties.tabColor = 'C9A84C'
ws_ps.freeze_panes = 'A5'

ws_ps.merge_cells('A1:J1')
c = ws_ps['A1']
c.value = 'SUNFORD VALVES — PRICE RANGE SUMMARY BY PRODUCT TYPE'
c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=13)
c.fill = PatternFill('solid', start_color=CLR['navy'])
c.alignment = Alignment(horizontal='center', vertical='center')
ws_ps.row_dimensions[1].height = 26

ws_ps.merge_cells('A2:J2')
c = ws_ps['A2']
c.value = 'Live-linked from 📋 Product Manager. Refresh with Ctrl+Alt+F9 to force recalculate. All prices INR Ex-Works Surat.'
c.font = Font(name='Arial', italic=True, color='FFB0C4DE', size=8)
c.fill = PatternFill('solid', start_color='FF0D2137')
c.alignment = Alignment(horizontal='center', vertical='center')
ws_ps.row_dimensions[2].height = 14

ws_ps.row_dimensions[3].height = 6

# Summary headers
sum_hdrs = ['Valve Type','Total SKUs','Active SKUs','Size Variants',
            'Min Price (₹)','Max Price (₹)','Avg Price (₹)',
            '"On Request" Items','Price Band','Sheet Link']
sum_widths = {1:20,2:11,3:11,4:13,5:15,6:15,7:15,8:17,9:18,10:14}
set_widths(ws_ps, sum_widths)

for ci, h in enumerate(sum_hdrs, 1):
    c = ws_ps.cell(row=4, column=ci, value=h)
    c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=9)
    c.fill = PatternFill('solid', start_color=CLR['navy'])
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin()
ws_ps.row_dimensions[4].height = 30

# For each valve type, build COUNTIF/MINIFS/MAXIFS/AVERAGEIFS formulas
valve_types_ordered = list(df['Valve_Type'].unique())
PM_RANGE_VT   = f"'📋 Product Manager'!$A${PM_DATA_START}:$A${LAST_DATA_ROW}"
PM_RANGE_RATE = f"'📋 Product Manager'!$K${PM_DATA_START}:$K${LAST_DATA_ROW}"
PM_RANGE_STS  = f"'📋 Product Manager'!$L${PM_DATA_START}:$L${LAST_DATA_ROW}"
PM_RANGE_SIZE = f"'📋 Product Manager'!$E${PM_DATA_START}:$E${LAST_DATA_ROW}"

for i, vtype in enumerate(valve_types_ordered):
    r = 5 + i
    bg = CLR['white'] if i % 2 == 0 else CLR['gray_lt']
    tab_c = TAB_COLORS.get(vtype, CLR['steel'])

    # Col A: Type name
    c = ws_ps.cell(row=r, column=1, value=vtype)
    c.font = Font(name='Arial', bold=True, size=9, color=tab_c[2:])
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = thin()

    # Col B: Total SKUs = COUNTIF
    c = ws_ps.cell(row=r, column=2, value=f'=COUNTIF({PM_RANGE_VT},A{r})')
    c.font = Font(name='Arial', size=9)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin()

    # Col C: Active SKUs = COUNTIFS
    c = ws_ps.cell(row=r, column=3,
        value=f'=COUNTIFS({PM_RANGE_VT},A{r},{PM_RANGE_STS},"Active")')
    c.font = Font(name='Arial', size=9, color=CLR['green'])
    c.fill = PatternFill('solid', start_color=CLR['green_lt'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin()

    # Col D: Unique sizes (can't auto-count unique easily with basic formulas, use SUMPRODUCT)
    c = ws_ps.cell(row=r, column=4,
        value=f'=SUMPRODUCT((({PM_RANGE_VT}=A{r})*1)/COUNTIFS({PM_RANGE_VT},{PM_RANGE_VT},{PM_RANGE_SIZE},{PM_RANGE_SIZE}),({PM_RANGE_VT}=A{r})*1)')
    c.font = Font(name='Arial', size=9)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin()

    # Col E: Min Price — MINIFS on numeric rates only
    c = ws_ps.cell(row=r, column=5,
        value=f'=IFERROR(MINIFS({PM_RANGE_RATE},{PM_RANGE_VT},A{r},{PM_RANGE_RATE},">0"),"-")')
    c.font = Font(name='Arial', size=9)
    c.fill = PatternFill('solid', start_color=CLR['gold_lt'])
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '₹\\ #,##0'
    c.border = thin()

    # Col F: Max Price
    c = ws_ps.cell(row=r, column=6,
        value=f'=IFERROR(MAXIFS({PM_RANGE_RATE},{PM_RANGE_VT},A{r},{PM_RANGE_RATE},">0"),"-")')
    c.font = Font(name='Arial', size=9)
    c.fill = PatternFill('solid', start_color=CLR['gold_lt'])
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '₹\\ #,##0'
    c.border = thin()

    # Col G: Avg Price = AVERAGEIFS
    c = ws_ps.cell(row=r, column=7,
        value=f'=IFERROR(AVERAGEIFS({PM_RANGE_RATE},{PM_RANGE_VT},A{r},{PM_RANGE_RATE},">0"),"-")')
    c.font = Font(name='Arial', bold=True, size=9)
    c.fill = PatternFill('solid', start_color=CLR['gold_lt'])
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '₹\\ #,##0'
    c.border = thin()

    # Col H: On Request = COUNTIFS rate = "On Request"
    c = ws_ps.cell(row=r, column=8,
        value=f'=COUNTIFS({PM_RANGE_VT},A{r},{PM_RANGE_RATE},"On Request")')
    c.font = Font(name='Arial', size=9, color=CLR['on_req'][2:])
    c.fill = PatternFill('solid', start_color=CLR['on_req_lt'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin()

    # Col I: Price Band
    c = ws_ps.cell(row=r, column=9,
        value=f'=IF(ISNUMBER(G{r}),IF(G{r}<5000,"⭐ Budget",IF(G{r}<20000,"⭐⭐ Mid",IF(G{r}<60000,"⭐⭐⭐ Premium","⭐⭐⭐⭐ High-End"))),"Custom")')
    c.font = Font(name='Arial', bold=True, size=8)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin()

    # Col J: Hyperlink to type sheet
    c = ws_ps.cell(row=r, column=10, value=f'Go to {vtype}')
    c.hyperlink = f'#{vtype}!A1'
    c.font = Font(name='Arial', size=9, color='FF1565C0', underline='single')
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin()

    ws_ps.row_dimensions[r].height = 18

# Totals row
tr = 5 + len(valve_types_ordered)
ws_ps.merge_cells(f'A{tr}:D{tr}')
c = ws_ps.cell(row=tr, column=1, value='PORTFOLIO TOTAL')
c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=9)
c.fill = PatternFill('solid', start_color=CLR['navy'])
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = thin()

total_data = [
    None,
    f'=MINIFS({PM_RANGE_RATE},{PM_RANGE_RATE},">0")',
    f'=MAXIFS({PM_RANGE_RATE},{PM_RANGE_RATE},">0")',
    f'=AVERAGEIFS({PM_RANGE_RATE},{PM_RANGE_RATE},">0")',
    f'=COUNTIF({PM_RANGE_RATE},"On Request")',
    f'=COUNTA({PM_RANGE_VT})',
    f'Go to Product Manager',
]
for ci, val in zip(range(5,11), total_data[1:]):
    c = ws_ps.cell(row=tr, column=ci, value=val)
    c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=9)
    c.fill = PatternFill('solid', start_color=CLR['navy'])
    c.alignment = Alignment(horizontal='center' if ci != 5 else 'right', vertical='center')
    c.border = thin()
    if ci in (5,6,7): c.number_format = '₹\\ #,##0'
ws_ps.row_dimensions[tr].height = 20

ws_ps.page_setup.orientation = 'landscape'
ws_ps.page_setup.fitToPage = True
ws_ps.print_title_rows = '1:4'

print("✓ Price Summary sheet built")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3–N: TYPE SHEETS — fully formula-linked to Product Manager
# ══════════════════════════════════════════════════════════════════════════════
# Build index: valve_type -> list of PM row numbers (1-indexed Excel rows)
type_to_pm_rows = {}
for vtype in df['Valve_Type'].unique():
    idxs = df.index[df['Valve_Type'] == vtype].tolist()
    type_to_pm_rows[vtype] = [PM_DATA_START + i for i in idxs]

def build_type_sheet(vtype, pm_rows):
    tab_c = TAB_COLORS.get(vtype, 'FF1A3A5C')
    ws = wb.create_sheet(title=vtype)
    ws.sheet_properties.tabColor = tab_c[2:]
    ws.freeze_panes = 'A5'

    col_widths = {1:5, 2:52, 3:32, 4:10, 5:14, 6:18, 7:32, 8:22, 9:7, 10:14, 11:9, 12:20}
    set_widths(ws, col_widths)

    # Row 1: Title
    ws.merge_cells('A1:L1')
    c = ws['A1']
    c.value = f'SUNFORD VALVES COMPANY  ·  {vtype.upper()}  ·  PRICE LIST'
    c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=13)
    c.fill = PatternFill('solid', start_color=CLR['navy'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Row 2: Live-link notice
    ws.merge_cells('A2:L2')
    c = ws['A2']
    c.value = '⚡ LIVE — All data below is formula-linked to 📋 Product Manager. Edit prices there, they update here automatically.'
    c.font = Font(name='Arial', italic=True, size=8, color='FFFFD080')
    c.fill = PatternFill('solid', start_color='FF0D2137')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 14

    # Row 3: spacer
    ws.row_dimensions[3].height = 4

    # Row 4: Column headers
    hdrs = ['Sr.','Product Description','MOC / Material','Size',
            'Pressure Class','End Connection','Design Standard',
            'Hydro Test','Unit','Rate (INR)','Status','Remarks']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=9)
        c.fill = PatternFill('solid', start_color=tab_c[2:])
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin()
    ws.row_dimensions[4].height = 32

    row = 5
    sr = 1

    df_type = df[df['Valve_Type'] == vtype].copy()
    df_type['__pm_row'] = pm_rows[:len(df_type)]  # ensure alignment

    prev_vname = None
    prev_moc = None

    for _, data_row in df_type.iterrows():
        pm_row = data_row['__pm_row']
        PM_REF = f"'📋 Product Manager'"

        vname = data_row['Valve_Name']
        moc = data_row['MOC']

        # Product group header (new valve_name)
        if vname != prev_vname:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            c = ws.cell(row=row, column=1)
            c.value = f"={PM_REF}!B{pm_row}"
            c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=8)
            c.fill = PatternFill('solid', start_color='FF1A3A5C')
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c.border = thin()
            ws.row_dimensions[row].height = 22
            row += 1

            # Design/hydro info row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            c = ws.cell(row=row, column=1)
            c.value = f'=IFERROR("  Design: "&{PM_REF}!H{pm_row}&"   ·   Hydro: "&{PM_REF}!I{pm_row}&" Kg/Cm²", "")'
            c.font = Font(name='Arial', italic=True, color='FF3C5C7A', size=8)
            c.fill = PatternFill('solid', start_color=CLR['sky'])
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.border = thin()
            ws.row_dimensions[row].height = 14
            row += 1

            prev_moc = None  # reset moc tracking

        # MOC sub-header (new moc within same vname)
        if moc != prev_moc:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            c = ws.cell(row=row, column=1)
            c.value = f'=IFERROR("  MOC: "&{PM_REF}!D{pm_row}, "")'
            c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=8)
            c.fill = PatternFill('solid', start_color='FF2B5F8C')
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.border = thin()
            ws.row_dimensions[row].height = 16
            row += 1

        bg_idx = (sr - 1) % 2
        bg = CLR['white'] if bg_idx == 0 else CLR['gray_lt']

        # Sr
        c = ws.cell(row=row, column=1, value=sr)
        c.font = Font(name='Arial', color='FF555555', size=8)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin()

        # Product Description = PM column B
        c = ws.cell(row=row, column=2, value=f"={PM_REF}!B{pm_row}")
        c.font = Font(name='Arial', size=8)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = thin()

        # MOC = PM column D
        c = ws.cell(row=row, column=3, value=f"={PM_REF}!D{pm_row}")
        c.font = Font(name='Arial', size=8)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = thin()

        # Size = PM column E
        c = ws.cell(row=row, column=4, value=f"={PM_REF}!E{pm_row}")
        c.font = Font(name='Arial', bold=True, size=9, color='FF1A3A5C')
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin()

        # Pressure Class = PM col F (fallback to G if blank)
        c = ws.cell(row=row, column=5, value=f'=IFERROR(IF({PM_REF}!F{pm_row}<>"",{PM_REF}!F{pm_row},{PM_REF}!G{pm_row}),"")')
        c.font = Font(name='Arial', size=8)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin()

        # End Connection = PM col G
        c = ws.cell(row=row, column=6, value=f"={PM_REF}!G{pm_row}")
        c.font = Font(name='Arial', size=8)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin()

        # Design Standard = PM col H
        c = ws.cell(row=row, column=7, value=f"={PM_REF}!H{pm_row}")
        c.font = Font(name='Arial', size=8)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = thin()

        # Hydro Test = PM col I
        c = ws.cell(row=row, column=8, value=f"={PM_REF}!I{pm_row}")
        c.font = Font(name='Arial', size=8)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin()

        # Unit = PM col J
        c = ws.cell(row=row, column=9, value=f"={PM_REF}!J{pm_row}")
        c.font = Font(name='Arial', size=8)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin()

        # Rate = PM col K
        c = ws.cell(row=row, column=10, value=f"={PM_REF}!K{pm_row}")
        c.font = Font(name='Arial', bold=True, size=9)
        c.fill = PatternFill('solid', start_color=CLR['gold_lt'])
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.number_format = '₹\\ #,##0;₹\\ #,##0;"On Request"'
        c.border = thin()

        # Status = PM col L
        c = ws.cell(row=row, column=11, value=f"={PM_REF}!L{pm_row}")
        c.font = Font(name='Arial', bold=True, size=8)
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin()

        # Remarks = PM col N
        c = ws.cell(row=row, column=12, value=f'=IFERROR(IF({PM_REF}!N{pm_row}="","",{PM_REF}!N{pm_row}),"")')
        c.font = Font(name='Arial', size=8, italic=True, color='FF666666')
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = thin()

        ws.row_dimensions[row].height = 15
        prev_vname = vname
        prev_moc = moc
        sr += 1
        row += 1

    # Back link
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value='← Back to Product Manager')
    c.hyperlink = '#\'📋 Product Manager\'!A1'
    c.font = Font(name='Arial', size=9, color='FF1565C0', underline='single')
    c.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    c = ws.cell(row=row, column=5, value='← Price Summary')
    c.hyperlink = '#\'📊 Price Summary\'!A1'
    c.font = Font(name='Arial', size=9, color='FF1565C0', underline='single')
    c.alignment = Alignment(horizontal='left', vertical='center')

    ws.print_title_rows = '1:4'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    return ws

for vtype in valve_types_ordered:
    pm_rows = type_to_pm_rows[vtype]
    build_type_sheet(vtype, pm_rows)
    print(f"  ✓ {vtype}")

print("✓ All type sheets built with live formulas")

# ══════════════════════════════════════════════════════════════════════════════
# HOW TO USE SHEET
# ══════════════════════════════════════════════════════════════════════════════
ws_h = wb.create_sheet('ℹ️ How to Use', 2)
ws_h.sheet_properties.tabColor = '555555'

ws_h.merge_cells('A1:G1')
ws_h['A1'].value = 'SFVC PRODUCT WORKBOOK — USER GUIDE'
ws_h['A1'].font = Font(name='Arial', bold=True, color='FFFFFFFF', size=12)
ws_h['A1'].fill = PatternFill('solid', start_color=CLR['navy'])
ws_h['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_h.row_dimensions[1].height = 24

set_widths(ws_h, {1:26,2:70,3:10,4:10,5:10,6:10,7:10})

guide = [
    ('SHEET OVERVIEW', '', True),
    ('📋 Product Manager', 'THE ONLY SHEET YOU SHOULD EDIT. All other sheets read from here via formulas. Changing a rate here = it instantly changes in the type sheet and price summary.', False),
    ('📊 Price Summary', 'Auto-calculated dashboard. COUNTIF/MINIFS/MAXIFS formulas linked to Product Manager. Refreshes on file open or Ctrl+Alt+F9.', False),
    ('Ball Valves, Gate Valves, etc.', 'Print-ready price sheets. Every cell is a formula pointing to Product Manager. Grouped by product description and MOC.', False),
    ('','', False),
    ('HOW LINKING WORKS', '', True),
    ('Formula Example', "Cell K8 in Product Manager holds ₹5,755. Ball Valves sheet cell shows =\\'📋 Product Manager\\'!K8 — it always reads K8.", False),
    ('Adding a product', 'Add a row in Product Manager. Then go to the relevant type sheet and add a new formula row pointing to the new PM row. (Type sheets have fixed row refs, not dynamic lookup.)', False),
    ('Price edit', 'Edit column K (Rate_INR) in Product Manager. All linked type sheets update immediately.', False),
    ('Status edit', 'Edit column L (Status) to "Inactive". The value shows on type sheets. Filter by Status="Active" for website use.', False),
    ('','', False),
    ('FOR WEBSITE / APP USE', '', True),
    ('JSON Export', 'The website should read 📋 Product Manager sheet directly (columns A–N, rows 8 onward). Filter out Status="Inactive" rows. Rate column is numeric for priced items, text "On Request" otherwise.', False),
    ('Column Map', 'A=Valve_Type, B=Valve_Name, C=Description, D=MOC, E=Size, F=Pressure_Class, G=End_Connection, H=Design_Standard, I=Hydro_Test, J=Unit, K=Rate_INR, L=Status, N=Remarks', False),
    ('','', False),
    ('COLOR LEGEND', '', True),
    ('Gold background', 'Priced item (numeric Rate)', False),
    ('Purple/mauve', '"On Request" item', False),
    ('Green', 'Active status', False),
    ('Red', 'Inactive status', False),
    ('Dark navy row', 'Product group header (pulled from PM)', False),
    ('Steel blue row', 'MOC sub-group header (pulled from PM)', False),
]

for i, (col1, col2, is_header) in enumerate(guide):
    r = i + 2
    if is_header:
        ws_h.merge_cells(f'A{r}:G{r}')
        c = ws_h.cell(row=r, column=1, value=col1)
        c.font = Font(name='Arial', bold=True, color='FFFFFFFF', size=9)
        c.fill = PatternFill('solid', start_color=CLR['steel'])
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = thin()
        ws_h.row_dimensions[r].height = 18
    elif col1 == '':
        ws_h.row_dimensions[r].height = 8
    else:
        c1 = ws_h.cell(row=r, column=1, value=col1)
        c1.font = Font(name='Arial', bold=True, size=9, color='FF1A3A5C')
        c1.fill = PatternFill('solid', start_color=CLR['sky'])
        c1.alignment = Alignment(horizontal='left', vertical='center')
        c1.border = thin()
        ws_h.merge_cells(f'B{r}:G{r}')
        c2 = ws_h.cell(row=r, column=2, value=col2)
        c2.font = Font(name='Arial', size=8)
        c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c2.border = thin()
        ws_h.row_dimensions[r].height = 16

print("✓ How to Use sheet built")

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(output_xlsx)
print(f"\n✅ Saved: {output_xlsx}")
